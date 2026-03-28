from __future__ import annotations

import sqlite3
from collections import defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import records_db

HEADERS = ["序号", "提交用户", "商品名", "数量", "单价", "总价", "提交时间"]
SHEET_RECORDS = "对账记录"
COLUMN_WIDTHS = (8, 14, 28, 10, 12, 12, 22)

_I_SUBMITTER = 1
_I_PRODUCT = 2
_I_QTY = 3
_I_TOTAL = 5


def _to_decimal(raw: object) -> Decimal | None:
    if raw is None:
        return None
    try:
        return Decimal(str(raw).strip())
    except (InvalidOperation, ValueError, TypeError):
        return None


def _header_row_matches(first: tuple[object, ...]) -> bool:
    if len(first) < len(HEADERS):
        return False
    for i, expected in enumerate(HEADERS):
        raw = first[i]
        if raw is None:
            return False
        if str(raw).strip() != expected:
            return False
    return True


def _is_blank_row(values: tuple[object, ...]) -> bool:
    for v in values:
        if v is None:
            continue
        if isinstance(v, str):
            if v.strip():
                return False
        elif str(v).strip():
            return False
    return True


def _submitted_at_str(raw: object) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.replace(microsecond=0).isoformat(sep=" ")
    s = str(raw).strip()
    return s if s else None


def _money_db_str(raw: object) -> str | None:
    d = _to_decimal(raw)
    if d is None or d < 0:
        return None
    q = d.quantize(Decimal("0.01"))
    return format(q, "f")


def _quantity_db_str(raw: object) -> str | None:
    d = _to_decimal(raw)
    if d is None or d < 0:
        return None
    s = format(d, "f")
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    return s


def _parse_import_row(values: tuple[object, ...]) -> tuple[str, str, str, str, str, str] | str:
    cells = list(values[: len(HEADERS)])
    while len(cells) < len(HEADERS):
        cells.append(None)
    _id_ignore, sub_by, prod, qty_raw, up_raw, tot_raw, at_raw = cells

    prod_s = (str(prod).strip() if prod is not None else "")
    if not prod_s:
        return "商品名不能为空。"

    qty_s = _quantity_db_str(qty_raw)
    if qty_s is None:
        return "数量无效或不能为负数。"

    up_s = _money_db_str(up_raw)
    if up_s is None:
        return "单价无效或不能为负数。"

    tot_s = _money_db_str(tot_raw)
    if tot_s is None:
        return "总价无效或不能为负数。"

    at_s = _submitted_at_str(at_raw)
    if at_s is None:
        return "提交时间不能为空。"

    by_s = str(sub_by).strip() if sub_by is not None else ""

    return (prod_s, qty_s, up_s, tot_s, at_s, by_s)


def merge_xlsx_into_database(db_path: Path, xlsx_path: Path) -> tuple[int, list[str]]:
    """Import rows from an exported workbook into the local database (new ids).

    Returns (inserted_count, error_messages) where errors refer to sheet row numbers.
    """
    records_db.ensure_schema(db_path)
    wb = None
    try:
        try:
            wb = load_workbook(xlsx_path, read_only=False, data_only=True)
        except OSError:
            raise
        except Exception as e:
            raise ValueError(f"无法打开 Excel 文件：{e}") from e

        if not wb.worksheets:
            raise ValueError("工作簿中没有任何工作表。")
        if SHEET_RECORDS in wb.sheetnames:
            ws = wb[SHEET_RECORDS]
        else:
            ws = wb.worksheets[0]

        first = next(
            ws.iter_rows(min_row=1, max_row=1, max_col=len(HEADERS), values_only=True),
            None,
        )
        if first is None:
            raise ValueError("工作表为空，请使用本工具导出的 Excel。")
        if not _header_row_matches(first):
            raise ValueError(
                "表头与导出格式不一致，请使用本工具生成的「对账记录」工作表。"
            )

        to_insert: list[tuple[str, str, str, str, str, str]] = []
        errors: list[str] = []

        for excel_row_idx, row_tuple in enumerate(
            ws.iter_rows(min_row=2, max_col=len(HEADERS), values_only=True),
            start=2,
        ):
            if _is_blank_row(row_tuple):
                continue
            parsed = _parse_import_row(row_tuple)
            if isinstance(parsed, str):
                errors.append(f"第{excel_row_idx}行：{parsed}")
            else:
                to_insert.append(parsed)

        if to_insert:
            records_db.insert_line_items_batch(db_path, to_insert)
        return len(to_insert), errors
    finally:
        if wb is not None:
            wb.close()


def _write_analysis_sheet(wb: Workbook, rows: list[tuple]) -> None:
    ws = wb.create_sheet("汇总分析", 1)
    bold = Font(bold=True)
    r = 1

    def section_title(text: str) -> None:
        nonlocal r
        ws.cell(row=r, column=1, value=text).font = bold
        r += 1

    def header_row(values: list[str]) -> None:
        nonlocal r
        for c, v in enumerate(values, start=1):
            ws.cell(row=r, column=c, value=v).font = bold
        r += 1

    total_amount = Decimal("0")
    for row in rows:
        d = _to_decimal(row[_I_TOTAL])
        if d is not None:
            total_amount += d

    section_title("一、总体")
    header_row(["指标", "数值"])
    ws.cell(row=r, column=1, value="明细条数")
    ws.cell(row=r, column=2, value=len(rows))
    r += 1
    ws.cell(row=r, column=1, value="金额合计")
    ws.cell(row=r, column=2, value=float(total_amount))
    ws.cell(row=r, column=2).number_format = "0.00"
    r += 1
    r += 1

    by_user: dict[str, list[int | Decimal]] = defaultdict(
        lambda: [0, Decimal("0")]
    )
    for row in rows:
        u = (row[_I_SUBMITTER] or "").strip() or "（未记录）"
        by_user[u][0] += 1
        d = _to_decimal(row[_I_TOTAL])
        if d is not None:
            by_user[u][1] += d

    section_title("二、按提交用户")
    header_row(["提交用户", "明细条数", "金额合计"])
    for u in sorted(by_user.keys()):
        cnt, amt = by_user[u]
        ws.cell(row=r, column=1, value=u)
        ws.cell(row=r, column=2, value=cnt)
        ws.cell(row=r, column=3, value=float(amt))
        ws.cell(row=r, column=3).number_format = "0.00"
        r += 1
    r += 1

    by_product: dict[str, list[int | Decimal]] = defaultdict(
        lambda: [0, Decimal("0"), Decimal("0")]
    )
    for row in rows:
        p = (row[_I_PRODUCT] or "").strip() or "（未命名）"
        by_product[p][0] += 1
        q = _to_decimal(row[_I_QTY])
        if q is not None:
            by_product[p][1] += q
        d = _to_decimal(row[_I_TOTAL])
        if d is not None:
            by_product[p][2] += d

    section_title("三、按商品")
    header_row(["商品名", "明细条数", "数量合计", "金额合计"])
    for p in sorted(by_product.keys()):
        cnt, qty_sum, amt_sum = by_product[p]
        ws.cell(row=r, column=1, value=p)
        ws.cell(row=r, column=2, value=cnt)
        ws.cell(row=r, column=3, value=float(qty_sum))
        ws.cell(row=r, column=4, value=float(amt_sum))
        ws.cell(row=r, column=3).number_format = "0.000"
        ws.cell(row=r, column=4).number_format = "0.00"
        r += 1

    for idx, w in enumerate((18, 12, 14, 14), start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def export_database_to_xlsx(
    db_path: Path,
    xlsx_path: Path,
    *,
    viewer_username: str,
    is_admin: bool,
) -> int:
    records_db.ensure_schema(db_path)
    with sqlite3.connect(db_path) as conn:
        if is_admin:
            cur = conn.execute(
                """
                SELECT id, submitted_by, product_name, quantity, unit_price,
                       total_price, submitted_at
                FROM line_items
                ORDER BY id ASC
                """
            )
        else:
            cur = conn.execute(
                """
                SELECT id, submitted_by, product_name, quantity, unit_price,
                       total_price, submitted_at
                FROM line_items
                WHERE submitted_by = ?
                ORDER BY id ASC
                """,
                (viewer_username,),
            )
        rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_RECORDS
    ws.append(HEADERS)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    for row in rows:
        ws.append(list(row))

    for idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    _write_analysis_sheet(wb, list(rows))

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    return len(rows)