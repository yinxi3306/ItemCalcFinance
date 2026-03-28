from __future__ import annotations

import json
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import List


@dataclass(frozen=True)
class Item:
    name: str
    unit_price: Decimal


@dataclass(frozen=True)
class Category:
    name: str
    items: List[Item]


EXCEL_SHEET_NAME = "商品目录"
EXCEL_HEADERS = ("分类", "商品名", "单价")


def load_catalog_from_path(path: Path) -> List[Category]:
    """按扩展名加载 JSON 或 Excel（.xlsx / .xlsm）商品目录。"""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return load_catalog_from_excel(path)
    return load_catalog(path)


def load_catalog_from_excel(path: Path) -> List[Category]:
    """从 Excel 加载：工作表「商品目录」（若无则用第一张表），表头 分类/商品名/单价。"""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError(
            "读取 Excel 商品目录需要先安装 openpyxl：pip install openpyxl"
        ) from e

    if not path.is_file():
        raise FileNotFoundError(str(path))

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if EXCEL_SHEET_NAME in wb.sheetnames:
            ws = wb[EXCEL_SHEET_NAME]
        elif wb.worksheets:
            ws = wb.worksheets[0]
        else:
            raise ValueError("Excel 中没有任何工作表。")

        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None or len(header) < len(EXCEL_HEADERS):
            raise ValueError("表头无效：至少需要三列「分类」「商品名」「单价」。")
        got = tuple(
            str(header[i]).strip() if header[i] is not None else ""
            for i in range(len(EXCEL_HEADERS))
        )
        if got != EXCEL_HEADERS:
            raise ValueError(
                f"表头需为：{'、'.join(EXCEL_HEADERS)}，当前为：{'、'.join(got)}"
            )

        order: list[str] = []
        by_cat: dict[str, list[Item]] = {}
        row_num = 1

        for row in rows:
            row_num += 1
            if row is None:
                continue
            cells = list(row[: len(EXCEL_HEADERS)])
            while len(cells) < len(EXCEL_HEADERS):
                cells.append(None)
            c_raw, n_raw, p_raw = cells

            def _blank(v: object) -> bool:
                if v is None:
                    return True
                if isinstance(v, str) and not v.strip():
                    return True
                return False

            if _blank(c_raw) and _blank(n_raw) and _blank(p_raw):
                continue
            if _blank(c_raw):
                raise ValueError(f"第{row_num}行：分类不能为空。")
            cname = str(c_raw).strip()
            if _blank(n_raw):
                raise ValueError(f"第{row_num}行：商品名不能为空。")
            iname = str(n_raw).strip()
            if _blank(p_raw):
                raise ValueError(f"第{row_num}行：单价不能为空。")
            try:
                unit_price = Decimal(str(p_raw).strip())
            except (InvalidOperation, ValueError, TypeError) as e:
                raise ValueError(f"第{row_num}行：单价无法解析为数字。") from e
            if unit_price < 0:
                raise ValueError(f"第{row_num}行：单价不能为负。")

            if cname not in by_cat:
                order.append(cname)
                by_cat[cname] = []
            by_cat[cname].append(Item(name=iname, unit_price=unit_price))
    finally:
        wb.close()

    if not order:
        raise ValueError("Excel 中没有任何商品数据行。")

    return [Category(name=c, items=by_cat[c]) for c in order]


def load_catalog(path: Path) -> List[Category]:
    text = path.read_text(encoding="utf-8")
    data = json.loads(text)
    categories_raw = data.get("categories")
    if not isinstance(categories_raw, list):
        raise ValueError("JSON 根节点需包含 categories 数组")

    categories: List[Category] = []
    for cat in categories_raw:
        if not isinstance(cat, dict):
            raise ValueError("分类项格式错误")
        name = cat.get("name")
        items_raw = cat.get("items")
        if not isinstance(name, str) or not name.strip():
            raise ValueError("分类名称无效")
        if not isinstance(items_raw, list):
            raise ValueError(f"分类「{name}」的 items 需为数组")

        items: List[Item] = []
        for it in items_raw:
            if not isinstance(it, dict):
                raise ValueError(f"分类「{name}」中存在无效商品项")
            iname = it.get("name")
            price = it.get("unit_price")
            if not isinstance(iname, str) or not iname.strip():
                raise ValueError(f"分类「{name}」中存在无效商品名称")
            if price is None:
                raise ValueError(f"商品「{iname}」缺少 unit_price")
            try:
                unit_price = Decimal(str(price))
            except Exception as exc:  # noqa: BLE001
                raise ValueError(f"商品「{iname}」单价无法解析为数字") from exc
            if unit_price < 0:
                raise ValueError(f"商品「{iname}」单价不能为负")
            items.append(Item(name=iname.strip(), unit_price=unit_price))

        categories.append(Category(name=name.strip(), items=items))

    if not categories:
        raise ValueError("至少需要一个商品分类")
    return categories
